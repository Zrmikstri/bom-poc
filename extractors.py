"""
extractors.py
Parsing logic lifted unchanged from update_dinh_muc.py (the original script).
Reads one DM Excel file -> list of BOM-line dicts. No writing/output logic here.
"""
import os, re
import openpyxl
import xlrd
from datetime import datetime




def parse_file_meta(fname):
    """Tra ve (base_name, date_obj, date_str, ver_str) tu ten file DM."""
    base = re.sub(r'\.(xlsx|xls)$', '', fname, flags=re.IGNORECASE)
    m_date = re.search(r'\.(\d{6})$', base)
    date_str, date_obj = '', None
    if m_date:
        date_str = m_date.group(1)
        try:
            date_obj = datetime.strptime(date_str, '%d%m%y')
        except Exception:
            pass
        base = base[:m_date.start()]
    m_ver = re.search(r'\.?(V\d*)\s*$', base, re.IGNORECASE)
    ver_str = ''
    if m_ver:
        ver_str = m_ver.group(1).upper()
        base = base[:m_ver.start()]
    # Bo prefix DM / DM TH (co ca unicode)
    base_norm = re.sub(r'^[ĐD]M( TH)? - ', '', base).strip()
    return base_norm, date_obj, date_str, ver_str


def ver_num(ver_str):
    """Tra ve so version de sort: V2→2, V1→1, V→0, ''→-1"""
    if not ver_str: return -1
    m = re.search(r'\d+', ver_str)
    return int(m.group()) if m else 0


def format_date_str(date_str):
    try:
        return datetime.strptime(date_str, '%d%m%y').strftime('%d/%m/%y')
    except Exception:
        return date_str


def clean_str(val):
    if val is None: return ''
    s = str(val).strip()
    if s.lower() in ('none', 'nan', '#ref!'): return ''
    return s


def normalize_xuong(xuong):
    s = re.sub(r'^[IVX]+\.\s*', '', xuong).strip()
    return 'XƯỞNG ĐÓNG GÓI' if 'ĐÓNG GÓI' in s.upper() else s


def is_section_header(row):
    for cell in row[:5]:
        c = clean_str(cell)
        if 'XƯỞNG' in c.upper():
            return c
    return None


def extract_product_info(all_rows):
    ten_sp, ma_sp = '', ''
    for row in all_rows[:15]:
        row = list(row)
        while len(row) < 10: row.append(None)
        full = ' '.join(clean_str(x) for x in row)
        if 'Tên sản phẩm' in full or 'TÊN SẢN PHẨM' in full.upper():
            for idx in [2, 5, 6, 7]:
                val = clean_str(row[idx]) if idx < len(row) else ''
                if val and val.lower() not in ('', 'hình ảnh') and 'sản phẩm' not in val.lower():
                    ten_sp = val; break
        if 'Mã sản phẩm' in full or 'MÃ SẢN PHẨM' in full.upper():
            for idx in [2, 5, 6, 7]:
                val = clean_str(row[idx]) if idx < len(row) else ''
                if val and not any(b in val.lower() for b in ['mã', 'sản phẩm', '* ', 'số đơn']):
                    try: float(val); continue
                    except: pass
                    ma_sp = val; break
    return ten_sp, ma_sp


def fmt_num(v):
    if v is None: return None
    try:
        f = float(str(v).replace(',', '.'))
        return int(f) if f == int(f) else round(f, 6)
    except:
        s = str(v).strip()
        return s if s not in ('', 'None') else None


def num_to_str(v):
    if v is None or v == '' or v == 0: return ''
    if isinstance(v, int): return str(v)
    if isinstance(v, float): return str(int(v)) if v == int(v) else f'{v:g}'
    return str(v).strip()


def make_quy_cach(day, rong, cao, dai):
    parts = [p for p in [num_to_str(day), num_to_str(rong), num_to_str(cao), num_to_str(dai)] if p]
    return 'x'.join(parts)


def extract_file(filepath):
    fname = os.path.basename(filepath)
    # Route to dedicated extractor for XƯỞNG NHỰA files
    if 'DAY NHUA' in fname.upper():
        return extract_file_day_nhua(filepath)
    rows = []
    name_fb = re.sub(r'\.\d{6}\.(xlsx|xls)$', '', fname)
    name_fb = re.sub(r'^[ĐD]M( TH)? - ', '', name_fb).strip()
    all_rows = []
    try:
        if filepath.endswith('.xlsx'):
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            dm_th = next((s for s in wb.sheetnames if 'ĐM TH' in s), None)
            if not dm_th: wb.close(); return rows
            ws = wb[dm_th]
            all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
            wb.close()
        else:
            wb = xlrd.open_workbook(filepath)
            dm_th = next((s for s in wb.sheet_names() if 'ĐM TH' in s), None)
            if not dm_th: return rows
            ws = wb.sheet_by_name(dm_th)
            all_rows = [ws.row_values(i) for i in range(ws.nrows)]
    except Exception as e:
        print(f'  ⚠ {fname}: {e}')
        return rows

    ten_sp, ma_sp = extract_product_info(all_rows)
    if not ten_sp: ten_sp = name_fb
    current_xuong = ''
    skip_b = {'PHẦN SẮT', 'PHỤ KIỆN', 'PHẦN NHÔM', 'PHẦN PHỤ KIỆN', 'PHẦN INOX',
              'MÃ VẬT LIỆU', 'Mã vật liệu', 'STT', 'Stt'}
    skip_c = {'Tên nguyên vật liệu', 'TÊN NGUYÊN VẬT LIỆU', 'TÊN NVL', ''}

    for row in all_rows:
        row = list(row)
        while len(row) < 20: row.append(None)
        sec = is_section_header(row)
        if sec: current_xuong = normalize_xuong(sec); continue
        col_a, col_b, col_c = clean_str(row[0]), clean_str(row[1]), clean_str(row[2])
        if not col_c or col_c in skip_c: continue
        if col_b in skip_b: continue
        if col_a in ('STT', 'Stt', 'Hình ảnh'): continue
        if re.match(r'^[IVX]+/$', col_a): continue
        if col_c.upper() in ('PHẦN SẮT', 'PHỤ KIỆN', 'PHẦN NHÔM', 'PHẦN PHỤ KIỆN'): continue
        if not current_xuong: continue
        ma_vl = col_b if col_b and '#REF!' not in col_b else ''
        qc    = make_quy_cach(fmt_num(row[3]), fmt_num(row[4]), fmt_num(row[5]), fmt_num(row[6]))
        dvt, sl_cai, hao, tong, yc = (
            clean_str(row[7]), fmt_num(row[9]), fmt_num(row[11]),
            fmt_num(row[13]), clean_str(row[14]))
        if sl_cai is None and tong is None: continue
        key = '_'.join(filter(None, [ten_sp, ma_vl, col_c]))
        rows.append({'Ten_SP': ten_sp, 'Ma_SP': ma_sp, 'Xuong': current_xuong,
                     'Ma_VL': ma_vl, 'Ten_VL': col_c, 'Quy_cach': qc,
                     'DVT': dvt, 'SL_cai': sl_cai, 'Hao_hut': hao,
                     'Tong_SL': tong, 'YC_CL': yc, 'Key': key})
    return rows


# Col indices in DM DAY NHUA sheet
_DAY_NHUA_TYPES = {
    8:  'Dây đai',
    9:  'Dây quấn lót',
    10: 'Dây giăng sườn',
    11: 'Dây đan',
    12: 'Dây quấn chân',
}

def extract_file_day_nhua(filepath):
    rows = []
    fname = os.path.basename(filepath)
    try:
        wb = xlrd.open_workbook(filepath)
        sheet_name = next((s for s in wb.sheet_names() if 'DAY NHUA' in s.upper()), None)
        if not sheet_name: return rows
        ws = wb.sheet_by_name(sheet_name)
        all_rows = [ws.row_values(i) for i in range(ws.nrows)]
    except Exception as e:
        print(f'  ⚠ {fname}: {e}')
        return rows

    cur_ten_sp, cur_ma_sp, cur_bophan = '', '', ''
    for row in all_rows:
        row = list(row)
        while len(row) < 16: row.append(None)
        col_a = clean_str(row[0])
        ma_sp_raw  = clean_str(row[1])
        ten_sp_raw = clean_str(row[2])
        if ten_sp_raw and col_a not in ('', 'Stt', 'STT'):
            try:
                float(col_a)
                cur_ten_sp = ten_sp_raw
                cur_ma_sp  = ma_sp_raw
                cur_bophan = ''
            except:
                pass
        bophan_raw = clean_str(row[3])
        if bophan_raw:
            cur_bophan = bophan_raw

        ma_day = clean_str(row[5])
        if not ma_day or ma_day in ('Mã dây', 'Dây đan', ''):
            continue
        _FOOTER_KW = ('XÁC NHẬN', 'LẬP BẢNG', 'KIỂM TRA', 'BAN GIÁM ĐỐC', 'GĐ. ')
        if any(kw in ma_day.upper() for kw in _FOOTER_KW):
            continue
        qc_raw = clean_str(row[6])
        sl     = fmt_num(row[13])
        if not isinstance(sl, (int, float)):
            continue
        loai_day = ' / '.join(v for k, v in _DAY_NHUA_TYPES.items()
                               if clean_str(row[k]).lower() == 'x')
        if not loai_day:
            loai_day = ma_day
        if not cur_ten_sp:
            continue
        key = cur_ten_sp + ('_' + cur_bophan if cur_bophan else '')
        rows.append({
            'Ten_SP': cur_ten_sp, 'Ma_SP': cur_ma_sp, 'Xuong': 'XƯỞNG NHỰA',
            'Ma_VL': ma_day, 'Ten_VL': loai_day, 'Quy_cach': qc_raw,
            'DVT': 'Kg', 'SL_cai': sl, 'Hao_hut': None,
            'Tong_SL': fmt_num(row[15]) if len(row) > 15 else None, 'YC_CL': '',
            'Key': key,
        })
    return rows


